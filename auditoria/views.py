from django.shortcuts import render, redirect
from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from .models import Empresa
from .models import Control
from .models import Evaluacion
from .models import Control
import openpyxl
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table
from django.shortcuts import render
from .models import Empresa, Evaluacion
from collections import Counter
import json
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from collections import Counter
import json

@login_required
def menu(request):
    return render(request, 'menu.html')

@login_required
def crear_empresa(request):
    if request.method == 'POST':
        Empresa.objects.create(
            nombre=request.POST.get('nombre'),
            nit=request.POST.get('nit'),
            categoria=request.POST.get('categoria'),
            representante_legal=request.POST.get('representante_legal'),
            fecha_registro=request.POST.get('fecha_registro') or None
        )
        return redirect('lista_empresas')

    return render(request, 'crear_empresa.html')

@login_required
def lista_empresas(request):
    empresas = Empresa.objects.all()
    return render(request, 'lista_empresas.html', {'empresas': empresas})

@login_required
def seleccionar_rol(request, empresa_id):

    if request.method == 'POST':
        rol = request.POST.get('rol')

        # guardamos en sesión (igual que Java)
        request.session['rol'] = rol
        request.session['empresa_id'] = empresa_id

        return redirect('evaluacion', empresa_id=empresa_id)

    return render(request, 'seleccionar_rol.html', {
        'empresa_id': empresa_id
    })

@login_required
def evaluacion(request, empresa_id):

    rol = request.session.get('rol')
    empresa = Empresa.objects.get(id=empresa_id)
    controles = Control.objects.all()

    if request.method == 'POST':

        for c in controles:
            estado = request.POST.get(f'control_{c.id}')

            Evaluacion.objects.create(
                empresa=empresa,
                control=c,
                estado=estado
            )

        return redirect('evaluacion', empresa_id=empresa.id)

    return render(request, 'evaluacion.html', {
        'controles': controles,
        'rol': rol,
        'empresa': empresa
    })

@login_required
def reporte_excel(request, empresa_id):
    empresa = Empresa.objects.get(id=empresa_id)
    evaluaciones = Evaluacion.objects.filter(empresa=empresa)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte ISO 27001"

    ws['A1'] = "Empresa:"
    ws['B1'] = empresa.nombre

    ws['A2'] = "Fecha:"
    ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws.append([])
    ws.append(["Código", "Control", "Categoría", "Estado"])

    for e in evaluaciones:
        ws.append([
            e.control.codigo,
            e.control.nombre,
            e.control.categoria,
            e.estado
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_{empresa.nombre}.xlsx'

    wb.save(response)
    return response

@login_required
def reporte_pdf(request, empresa_id):
    empresa = Empresa.objects.get(id=empresa_id)
    evaluaciones = Evaluacion.objects.filter(empresa=empresa)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_{empresa.nombre}.pdf'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()

    elementos = []


    elementos.append(Paragraph(f"Reporte ISO 27001", styles['Title']))
    elementos.append(Spacer(1, 10))

    elementos.append(Paragraph(f"<b>Empresa:</b> {empresa.nombre}", styles['Normal']))
    elementos.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))


    elementos.append(Spacer(1, 20))

    data = [["Código", "Control", "Estado"]]

    for e in evaluaciones:
        data.append([
            e.control.codigo,
            e.control.nombre,
            e.estado
        ])

    tabla = Table(data)

    elementos.append(tabla)

    doc.build(elementos)

    return response

@login_required
def dashboard(request, empresa_id):

    empresa = Empresa.objects.get(id=empresa_id)
    evaluaciones = Evaluacion.objects.filter(empresa=empresa)

    estados = [e.estado for e in evaluaciones]
    conteo_estados = Counter(estados)

    cumple = conteo_estados.get("Cumple", 0)
    no_cumple = conteo_estados.get("NoCumple", 0)
    proceso = conteo_estados.get("Proceso", 0)

    total = evaluaciones.count()
    porcentaje = round((cumple / total) * 100, 2) if total > 0 else 0

    categorias = {}
    for e in evaluaciones:
        cat = e.control.categoria
        if cat not in categorias:
            categorias[cat] = {"Cumple":0, "NoCumple":0, "Proceso":0}
        
        categorias[cat][e.estado] += 1

    riesgo = {"Alto": 0, "Medio": 0, "Bajo": 0}

    for e in evaluaciones:
        if e.estado == "NoCumple":
            riesgo["Alto"] += 1
        elif e.estado == "Proceso":
            riesgo["Medio"] += 1
        else:
            riesgo["Bajo"] += 1

    madurez = round(
        ((cumple * 1) + (proceso * 0.5)) / total * 100, 2
    ) if total > 0 else 0

    cumplimiento_categoria = {}
    for cat, valores in categorias.items():
        total_cat = sum(valores.values())
        cumplimiento_categoria[cat] = round(
            (valores["Cumple"] / total_cat) * 100, 2
        ) if total_cat > 0 else 0

    ranking = sorted(
        cumplimiento_categoria.items(),
        key=lambda x: x[1],
        reverse=True
    )

    if porcentaje >= 80:
        color = "green"
    elif porcentaje >= 50:
        color = "orange"
    else:
        color = "red"

    return render(request, 'dashboard.html', {
        'empresa': empresa,
        'cumple': cumple,
        'no_cumple': no_cumple,
        'proceso': proceso,
        'porcentaje': porcentaje,
        'categorias': json.dumps(categorias),
        'riesgo': json.dumps(riesgo),
        'madurez': madurez,
        'ranking': ranking,
        'color': color
    })

def login_view(request):

    if request.user.is_authenticated:
        return redirect('menu')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            next_url = request.GET.get('next')
            if next_url and next_url != '/':
                return redirect(next_url)

            return redirect('menu')

        else:
            return render(request, 'login.html', {
                'error': 'Usuario o contraseña incorrectos'
            })

    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def registro_view(request):

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm = request.POST.get('confirm')

        if password != confirm:
            return render(request, 'registro.html', {
                'error': 'Las contraseñas no coinciden'
            })

        if User.objects.filter(username=username).exists():
            return render(request, 'registro.html', {
                'error': 'El usuario ya existe'
            })

        user = User.objects.create(
            username=username,
            email=email,
            password=make_password(password)
        )

        return redirect('login')

    return render(request, 'registro.html')


from django.shortcuts import render, redirect
from django.contrib import messages
from .models import DocumentoSGSI
import json
from datetime import date

# Datos iniciales precargados (se insertan solo si la tabla está vacía)
DOCUMENTOS_INICIALES = [
    ('DG-001', 'DG', 'Manual del SGSI', '/docs/general'),
    ('CD-001', 'CD', 'Procedimiento de Control de Documentos y Registros', '/docs/control'),
    ('VR-001', 'VR', 'Metodología de Evaluación y Tratamiento de Riesgos', '/docs/riesgos'),
    ('CC-001', 'CC', 'Plan de Concientización en Seguridad', '/docs/concientizacion'),
    ('AI-001', 'AI', 'Procedimiento de Auditoría Interna', '/docs/auditoria'),
    ('AC-001', 'AC', 'Procedimiento de Acciones Correctivas y Preventivas', '/docs/mejora'),
    ('RD-001', 'RD', 'Procedimiento de Revisión por la Dirección', '/docs/direccion'),
    ('PS-001', 'PS', 'Política de Seguridad de la Información', '/docs/politicas'),
    ('OS-001', 'OS', 'Política de Roles y Responsabilidades de Seguridad', '/docs/organizacion'),
    ('RH-001', 'RH', 'Procedimiento de Seguridad en Recursos Humanos', '/docs/rrhh'),
    ('GA-001', 'GA', 'Inventario y Clasificación de Activos de Información', '/docs/activos'),
    ('CA-001', 'CA', 'Política de Control de Acceso', '/docs/acceso'),
    ('EN-001', 'EN', 'Política de Uso de Controles Criptográficos', '/docs/encriptacion'),
    ('SF-001', 'SF', 'Política de Seguridad Física y del Entorno', '/docs/fisica'),
    ('SO-001', 'SO', 'Procedimientos de Operación Documentados', '/docs/operacion'),
    ('SC-001', 'SC', 'Política de Seguridad en Redes y Comunicaciones', '/docs/comunicaciones'),
    ('AD-001', 'AD', 'Política de Seguridad en Desarrollo de Software', '/docs/desarrollo'),
    ('RP-001', 'RP', 'Política de Seguridad con Proveedores', '/docs/proveedores'),
    ('GI-001', 'GI', 'Procedimiento de Gestión de Incidentes de Seguridad', '/docs/incidentes'),
    ('CN-001', 'CN', 'Plan de Continuidad del Negocio', '/docs/continuidad'),
    ('CU-001', 'CU', 'Procedimiento de Verificación de Cumplimiento Legal', '/docs/cumplimiento'),
]

def lista_maestra(request):
    # Precarga documentos si la tabla está vacía
    if DocumentoSGSI.objects.count() == 0:
        for codigo, cat, nombre, ubic in DOCUMENTOS_INICIALES:
            DocumentoSGSI.objects.create(
                codigo=codigo,
                categoria=cat,
                nombre=nombre,
                ubicacion=ubic,
                version='1.0'
            )

    documentos = DocumentoSGSI.objects.all()
    categorias = DocumentoSGSI.CATEGORIAS
    return render(request, 'lista_maestra.html', {
        'documentos': documentos,
        'categorias': categorias,
    })

def actualizar_documento(request, doc_id):
    if request.method == 'POST':
        doc = DocumentoSGSI.objects.get(id=doc_id)
        doc.version = request.POST.get('version', doc.version)
        doc.nombre = request.POST.get('nombre', doc.nombre)
        doc.ubicacion = request.POST.get('ubicacion', doc.ubicacion)
        fecha = request.POST.get('ultima_revision')
        if fecha:
            doc.ultima_revision = fecha
        doc.save()
        messages.success(request, 'Documento actualizado correctamente.')
    return redirect('lista_maestra')